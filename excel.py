import xlrd
import xlwt
import xlutils
from xlutils.copy import copy
import xlsxwriter
import openpyxl
import xlwings as xw 
import os

class EXCEL(object):
    def __init__(self):
        self.__name__ = "EXCEL"        
        self.workbook = xlwt.Workbook()

    # def create_file(self, filename):
        
    def check_file(self, file_name):
        if os.path.exists(file_name) == True:
            os.remove(file_name)
            
        self.create_file(file_name)
            
    def create_file(self, file_name):
        workbook = xlsxwriter.Workbook(file_name) 
        sheetname = 'sheet1'
        worksheet = workbook.get_worksheet_by_name(sheetname)

        # if worksheet is None:
        #     worksheet = workbook.add_worksheet(sheetname)        

    def get_data_bysheet(self, filename, sheetname = u'Sheet1'):
        print ("filename: ", filename)
        self.readHandle =  xlrd.open_workbook(filename)
        table = self.readHandle.sheet_by_name(sheetname)
        secode_list = []
        for i in range(table.nrows):
            secode_list.append(str(table.cell(i,0).value))
        return secode_list

    def get_onecolumn_data_by_sheet(self, filename, sheetname = u'Sheet1', column=0):
        print ("filename: ", filename)
        self.readHandle =  xlrd.open_workbook(filename)
        table = self.readHandle.sheet_by_name(sheetname)
        result = []
        if column > table.ncols:
            return result

        for i in range(table.nrows):
            result.append(str(table.cell(i, column).value))
        return result

    def get_alldata_bysheet(self, filename, sheetname = u'Sheet1'):
        print ("filename: ", filename)
        self.readHandle =  xlrd.open_workbook(filename)
        if sheetname == 'all':
            sheet_name_list = self.readHandle.sheet_names()
            result = {}
            for sheet_name in sheet_name_list:
                curr_result = []
                table = self.readHandle.sheet_by_name(sheet_name)
                print("sheet_name: %s, rows: %d, cols: %d " % \
                     (sheet_name, table.nrows, table.ncols))
                for row in range(table.nrows):
                    row_data = []
                    for col in range(table.ncols):
                        row_data.append(table.cell(row, col).value)
                    curr_result.append(row_data)
                result[sheet_name] = curr_result
        else:
            table = self.readHandle.sheet_by_name(sheetname)
            result = []            
            row_index = 0
            while row_index < table.nrows:
                row_data = []
                col_index = 0
                while col_index < table.ncols:
                    row_data.append(table.cell(row_index, col_index).value)
                    col_index += 1
                row_index += 1
                result.append(row_data)
        return result        

    def get_data_byindex(self, filename, index = 0):
        self.readHandle =  xlrd.open_workbook(filename)
        table = self.readHandle.sheet_by_index(index)
        secode_list = []
        for i in range(table.nrows):
            secode_list.append(str(int(table.cell(i,0).value)))
        return secode_list

    def write_data(self, oridata, file_name, sheetname = u"Sheet1"):
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet(sheetname, cell_overwrite_ok=False)

        for i in range(len(oridata)):
            for j in range(len(oridata[i])):
                sheet.write(i, j, oridata[i][j])
        workbook.save(file_name)

    def init_xlwt(self, file_name):
        self.readbook = xlrd.open_workbook(file_name)        
        self.workbook = copy(self.readbook)

    def save_xlwt(self, file_name):
        self.workbook.save(file_name)

    def write_single_column_data_xlwt(self, oridata, file_name, col=0, sheetname = u"Sheet1"):
        readbook = xlrd.open_workbook(file_name)        
        workbook = copy(readbook)
        sheet_namelist = readbook.sheet_names()
        if sheetname not in sheet_namelist:
            worksheet = workbook.add_sheet(sheetname, cell_overwrite_ok=True)
        else:            
            worksheet = workbook.get_sheet(sheetname)

        for i in range(len(oridata)):
                worksheet.write(i, col, oridata[i])

        workbook.save(file_name)        

    def write_all_data_xlwt(self, oridata, file_name, sheetname = "Sheet1"):
        readbook = xlrd.open_workbook(file_name)        
        workbook = copy(readbook)
        if sheetname not in readbook.sheet_names():
            worksheet = workbook.add_sheet(sheetname, cell_overwrite_ok=True)
        else:            
            worksheet = workbook.get_sheet(sheetname)

        for i in range(len(oridata)):
            for j in range(len(oridata[i])):
                worksheet.write(i, j, oridata[i][j])

        workbook.save(file_name)

    def write_single_column_data_xlsxwriter(self, oridata, file_name, col=0, sheetname = "Sheet1"):
        workbook = xlsxwriter.Workbook(file_name) 
        worksheet = workbook.get_worksheet_by_name(sheetname)
        
        if worksheet is None:
            worksheet = workbook.add_worksheet(sheetname)

        for i in range(len(oridata)):
            worksheet.write(i, col, oridata[i])

        workbook.close()   

    def write_all_data_xlsxwriter_turnover(self, oridata, file_name, sheetname = "Sheet1"):
        workbook = xlsxwriter.Workbook(file_name) 
        worksheet = workbook.get_worksheet_by_name(sheetname)
        
        if worksheet is None:
            worksheet = workbook.add_worksheet(sheetname)

        
        for i in range(len(oridata)):
            for j in range(len(oridata[i])):
                worksheet.write(j, i, oridata[i][j])
            print ("i: ", i+1, ", secode: ", oridata[i][0], ", numb: ", len(oridata[i]))

        workbook.close()   

    # 每次都会新建文件，新建sheet.
    def write_all_data_xlsxwriter_ori(self, oridata, file_name, sheetname = "Sheet1"):
        workbook = xlsxwriter.Workbook(file_name) 
        worksheet = workbook.get_worksheet_by_name(sheetname)
        print('file_name: %s' % (file_name))
        
        if worksheet is None:
            worksheet = workbook.add_worksheet(sheetname)
        
        for i in range(len(oridata)):
            for j in range(len(oridata[i])):
                worksheet.write(i, j, oridata[i][j])
            # print ('write i : %d')
        workbook.close()   

    def write_single_column_data_openpyxl(self, oridata, file_name, col=0, sheetname = "Sheet1"):
        if os.path.exists(file_name):
            workbook = openpyxl.load_workbook(file_name)
        else:
            workbook = openpyxl.Workbook()
            for cur_sheetname in workbook.sheetnames:
                workbook.remove(workbook[cur_sheetname])

        if sheetname not in workbook.sheetnames:
            print ('creat worksheet: ', sheetname)
            worksheet = workbook.create_sheet(sheetname)
        else:
            worksheet = workbook[sheetname]    

        for i in range(len(oridata)):
            worksheet.cell(row=i+1, column=col+1, value=oridata[i])

        workbook.save(file_name) 

    # 不会每次新建文件
    def write_all_data_openpyxl(self, oridata, file_name, sheetname = "Sheet1"):
        if os.path.exists(file_name):
            workbook = openpyxl.load_workbook(file_name)
        else:
            workbook = openpyxl.Workbook()
            for cur_sheetname in workbook.sheetnames:
                workbook.remove(workbook[cur_sheetname])

        if sheetname not in workbook.sheetnames:
            # print ('creat worksheet: ', sheetname)
            worksheet = workbook.create_sheet(sheetname)
        else:
            worksheet = workbook[sheetname]    

        for i in range(len(oridata)):
            for j in range(len(oridata[i])):
                worksheet.cell(row=i+1, column=j+1, value=oridata[i][j])
            # print ("i: ", i+1, ", secode: ", oridata[i][0], ", numb: ", len(oridata[i]))
        workbook.save(file_name) 

    def write_all_data_xlwings(self, ori_data, file_name, sheetname = 'Sheet1'):
        app = xw.App(visible=True, add_book=False)
        app.display_alerts=False
        app.screen_updating=False
        wb = app.books.add()
        
        data_rows = len(ori_data)
        data_cols = len(ori_data[0])
        # wb.sheets[sheetname][0:data_rows, 0:data_cols] = ori_data
        # wb.sheets[sheetname].range('A1').options(expand='table') = ori_data
        if sheetname not in  wb.sheets:
            print (wb.sheets)
            wb.sheets.add(sheetname)
        wb.sheets[sheetname].range('A1').value = ori_data

        # row_index = 0        
        # while row_index < len(ori_data):
        #     col_index = 0
        #     while col_index < len(ori_data[row_index]):
        #         wb.sheets[sheetname][]
        #         col_index += 1
        #     row_index += 1
            
        wb.save(file_name)
        wb.close()
        app.quit()

    def write_single_column_data(self, oridata, file_name, col=0, sheetname = "Sheet1"):
        # self.write_single_column_data_openpyxl(oridata, file_name, col, sheetname)
        # self.write_single_column_data_xlwt(oridata, file_name, col, sheetname)
        self.write_single_column_data_xlsxwriter(oridata, file_name, col, sheetname)
        #self.write_single_column_data_xlwt(oridata, file_name, col, sheetname)

    def write_all_data(self, oridata, file_name, sheetname = "Sheet1",style="turnover"):
        # if style == 'turnover':
        #     # self.write_all_data_openpyxl(oridata, file_name, sheetname)
        #     self.write_all_data_xlsxwriter_turnover(oridata, file_name, sheetname)
        #     # self.write_all_data_xlwt(oridata, file_name, sheetname)
        
        # if style == 'ori':
        #     # self.write_all_data_xlsxwriter_ori(oridata, file_name, sheetname)
        #     self.write_all_data_openpyxl(oridata, file_name, sheetname)

        self.write_all_data_openpyxl(oridata, file_name, sheetname)
        # self.write_all_data_xlsxwriter_ori(oridata, file_name, sheetname)
        # self.write_all_data_xlwings(oridata, file_name, sheetname)
            
